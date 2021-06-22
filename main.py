import telebot
from selenium import webdriver
from time import sleep
from telebot import types
import openpyxl

# -----------------------------------------------
# -----------------initial_data:-----------------
# -----------------------------------------------

book = openpyxl.Workbook()
sheet = book.active
driver = webdriver.Firefox()
url = "https://www.google.ru/"
bot = telebot.TeleBot('1815430029:AAFSDKBXxJLmNigBBlVd5ont6NJdeAu2y48')
kategor_video = ""
kategor_canal = ''
count = 5
flag = False


# ------------------------------------------------
# --------------------/start:---------------------
# ------------------------------------------------


@bot.message_handler(commands=['start'])
def start(message):
    sti = open('C:/Users/admin/PycharmProjects/Bots/sticker.webp', 'rb')

    # keyboard
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    settings_button = types.KeyboardButton("Настройки")
    slovo_button = types.KeyboardButton("Ввести фразу")
    utuber_button = types.KeyboardButton("Ввести название канала")
    markup.add(settings_button, slovo_button, utuber_button)

    bot.send_sticker(message.chat.id, sti)
    mes = "  Я могу помочь тебе с составлением подборки видео на ютуб по ключевой фразе или по названию" \
          " канала. Ты можешь изменить критерий выбора или количество видео в подборке, для этого," \
          " перейди, в раздел настройки. Если же все готово, то можем начинать. Выбери способ поиска."
    bot.send_message(message.chat.id, mes, reply_markup=markup)

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# -----------------text_message:------------------
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


@bot.message_handler(content_types=['text'])
def lalala(message):
    if message.chat.type == 'private':
        global flag
        global count
        if flag:
            texxt = (message.text)
            if texxt.isdigit():
                if int(texxt) < 1 or int(texxt) > 10:
                    bot.send_message(message.chat.id, 'Число не может быть меньше 1 или больше 10, введите новое число')
                else:
                    count = int(texxt)
                    flag = False
                    bot.send_message(message.chat.id, 'Количество видео в подборке успешно изменено на '+str(count))
            else:
                bot.send_message(message.chat.id, 'То, что вы ввели, не является числом, повторите попытку')

        elif message.text == 'Ввести фразу':
            search_videos(message)
        elif message.text == 'Ввести название канала':
            search_channel(message)
        elif message.text == 'Настройки':
            markup = types.InlineKeyboardMarkup(row_width=3)
            kolvo = types.InlineKeyboardButton("Количество", callback_data='kolv')
            kategoria_video = types.InlineKeyboardButton("По фразе", callback_data='kat1')
            kategoria_kanal = types.InlineKeyboardButton("По каналу", callback_data='kat2')
            markup.add(kolvo, kategoria_video, kategoria_kanal)
            bot.send_message(message.chat.id,
                             'Выбери, что ты хочешь изменить: количество видео в подборке,'
                             'или методы сортировки для поиска по фразе или по названию канала', reply_markup=markup)
        else:
            bot.send_message(message.chat.id, 'Я не разговорчивый, выбери команду при помощи клавиатуры снизу :)')

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ------------------Callback:---------------------
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


@bot.callback_query_handler(func=lambda call: True)
def callbacl_inline(call):
    try:
        if call.message:
            if call.data == 'kolv':
                global count
                global flag
                flag = True
                bot.send_message(call.message.chat.id,
                                 'Введи количество видео, из которых будет состоять подборка (до 10).')
                bot.answer_callback_query(callback_query_id=call.id, show_alert=True,
                                          text="Сейчас количество видео в подборке равно " + str(count))
            elif call.data == 'kat1':
                markup = types.InlineKeyboardMarkup(row_width=3)
                data = types.InlineKeyboardButton('По дате', callback_data='data')
                prosmotr = types.InlineKeyboardButton('По просмотрам', callback_data='prosmotr')
                reiting = types.InlineKeyboardButton('По рейтингу', callback_data='reiting')
                markup.add(data, prosmotr, reiting)
                bot.send_message(call.message.chat.id, 'Выбери, по какому принципу мне сортировать видео',
                                 reply_markup=markup)
            elif call.data == 'data':
                global kategor_video
                kategor_video = "&sp=CAI%253D"
                bot.send_message(call.message.chat.id, 'Метод сортировки успешно изменен, теперь видео при поиске по фразе сортируются по дате')
            elif call.data == 'prosmotr':
                kategor_video
                kategor_video = "&sp=CAM%253D"
                bot.send_message(call.message.chat.id, 'Метод сортировки успешно изменен, теперь видео при поиске по фразе сортируются по количеству просмотров')
            elif call.data == 'reiting':
                kategor_video
                kategor_video = "&sp=CAE%253D"
                bot.send_message(call.message.chat.id, 'Метод сортировки успешно изменен, теперь видео при поиске по фразе сортируются по рейтингу')
            elif call.data == 'kat2':
                markup = types.InlineKeyboardMarkup(row_width=3)
                most_popular = types.InlineKeyboardButton('Популярность', callback_data='popular')
                new = types.InlineKeyboardButton('Сначала новые', callback_data='new')
                old = types.InlineKeyboardButton('Сначала старые', callback_data='old')
                markup.add(most_popular, new, old)
                bot.send_message(call.message.chat.id, 'Выбери, по какому принципу мне сортировать видео', reply_markup=markup)
            elif call.data == 'popular':
                global kategor_kanal
                kategor_kanal = "?view=0&sort=p&flow=grid"
                bot.send_message(call.message.chat.id, 'Метод сортировки успешно изменен, теперь видео при поиске по названию канала сортируются по популярности')
            elif call.data == 'new':
                kategor_kanal = "?view=0&sort=dd&flow=grid"
                bot.send_message(call.message.chat.id, 'Метод сортировки успешно изменен, теперь видео при поиске по названию канала сортируются по новизне')
            elif call.data == 'old':
                kategor_kanal = "?view=0&sort=da&flow=grid"
                bot.send_message(call.message.chat.id, 'Метод сортировки успешно изменен, теперь при поиске видео по названию канала сначала будут найдены старые видео')
            elif call.data == 'save':
                f = open('bot_result.xlsx', 'rb')
                bot.send_document(call.message.chat.id, f)

    except Exception as e:
        print(repr(e))
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ------------------Functions:--------------------
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

def search_videos(message):
    msg = bot.send_message(message.chat.id, "Введите фразу, по которой мне искать видео на YouTube")
    bot.register_next_step_handler(msg, search)


def search_channel(message):
    msg = bot.send_message(message.chat.id,
                           "Введите название YouTube канала (похожее на то, что на ютубе, иначе могу найти сайт Пм-Пу)")
    bot.register_next_step_handler(msg, search_from_channel)


def search_from_channel(message):
    global count
    bot.send_message(message.chat.id, "Сейчас соберу и вышлю тебе подборочку:)")
    driver.get(str(search_name(message.text)) + "/videos")
    videos = driver.find_elements_by_id("video-title")
    for i in range(count):
        bot.send_message(message.chat.id, videos[i].get_attribute('href'))
        info = videos[i].get_attribute('href') + ' ' + videos[i].get_attribute('aria-label')
        sheet.cell(row=2 * (i + 1), column=1).value = videos[i].get_attribute('href') + "  " * 20 + videos[
            i].get_attribute('aria-label')

        book.save("bot_result.xlsx")
    save = types.InlineKeyboardButton("Сохранить", callback_data='save')
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(save)
    bot.send_message(message.chat.id, "Если вы хотите получить более подробный результат в виде таблицы, нажмите 'Получить Excel'", reply_markup=markup)


def search_name(message):
    video_href = "https://www.youtube.com/results?search_query=" + message
    driver.get(video_href)
    sleep(2)
    name = driver.find_element_by_id("main-link")
    ssil = name.get_attribute('href')
    return ssil


def search(message):
    global count
    bot.send_message(message.chat.id, "Сейчас соберу и вышлю тебе подборочку")
    video_href = "https://www.youtube.com/results?search_query=" + message.text + kategor_video
    driver.get(video_href)
    sleep(2)
    videos = driver.find_elements_by_id("video-title")
    for i in range(count):
        bot.send_message(message.chat.id, videos[i].get_attribute('href'))
        info = videos[i].get_attribute('href') + ' ' + videos[i].get_attribute('aria-label')
        sheet.cell(row=2 * (i + 1), column=1).value = videos[i].get_attribute('href') + "  " * 20 + videos[i].get_attribute('aria-label')

        book.save("bot_result.xlsx")
    save = types.InlineKeyboardButton("Получить Excel", callback_data='save')
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(save)
    bot.send_message(message.chat.id, "Если вы хотите получить более подробный результат в виде таблицы, нажмите 'Получить'", reply_markup=markup)


bot.polling()
